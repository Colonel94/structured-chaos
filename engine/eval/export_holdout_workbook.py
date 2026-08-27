"""Export a labelling workbook's Cases sheet to the CSV the scorer consumes, and run QA on it.

`holdout_labels.xlsx` is the owner source of truth (Cases + Option Sets + QA Summary, dropdown-validated).
`score_holdout.py` reads `holdout_labels_<name>.csv`. This regenerates that CSV from a workbook so the two
never drift. Re-run after editing the workbook.

    cd engine && uv run --group dev python eval/export_holdout_workbook.py                 # -> holdout_labels_owner.csv
    cd engine && uv run --group dev python eval/export_holdout_workbook.py alice --from eval/fixtures/holdout_labels_alice_returned.xlsx
                                                                                            # -> holdout_labels_alice.csv (an independent labeller's returned file)
    cd engine && uv run --group dev python eval/export_holdout_workbook.py --qa             # integrity + full distributions of the default workbook
    cd engine && uv run --group dev python eval/export_holdout_workbook.py --qa --from <path>

`--from <path>` points at any returned workbook (same Cases/Option Sets layout) — this is how an
INDEPENDENT labeller's file becomes `holdout_labels_<name>.csv` (column 2 of the four numbers) without
overwriting the owner's. The --qa report covers what the workbook's QA Summary omits (owner review
2026-08-26): severity/emotion distributions, missing-label counts per field, duplicate ids/narratives,
source composition, and an out-of-vocabulary check of every gold cell against the Option Sets sheet.
"""

from __future__ import annotations

import collections
import csv
import sys
from pathlib import Path

import openpyxl

_DIR = Path(__file__).resolve().parent / "fixtures"
_WORKBOOK = _DIR / "holdout_labels.xlsx"
_GOLD = (
    "gold_category",
    "gold_desired_outcome",
    "gold_severity_signal",
    "gold_emotion_signal",
)


def _cases(workbook: Path) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    rows = list(wb["Cases"].iter_rows(values_only=True))
    header = [(h or "").strip() for h in rows[0]]
    data = [r for r in rows[1:] if r is not None and r[0] not in (None, "")]
    return header, data


def _option_sets(workbook: Path) -> dict[str, set[str]]:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    allowed: dict[str, set[str]] = collections.defaultdict(set)
    for r in list(wb["Option Sets"].iter_rows(values_only=True))[1:]:
        if r and r[0] and r[1]:
            allowed[str(r[0]).strip()].add(str(r[1]).strip())
    return allowed


def export(name: str, workbook: Path) -> None:
    header, data = _cases(workbook)
    out = _DIR / f"holdout_labels_{name}.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in data:
            w.writerow(["" if c is None else str(c) for c in r])
    print(f"wrote {out} — {len(data)} rows + header  (from {workbook.name})")


def qa(workbook: Path) -> int:
    header, data = _cases(workbook)
    idx = {h: i for i, h in enumerate(header)}
    allowed = _option_sets(workbook)
    print(f"QA — {workbook.name} — {len(data)} cases\n" + "=" * 60)

    # source composition
    src: collections.Counter[str] = collections.Counter()
    for r in data:
        i = str(r[idx["id"]])
        src["CFPB" if i.isdigit() else i.split("-")[0]] += 1
    print("source:", dict(src))

    # distributions + missing counts for all four gold fields
    for col in _GOLD:
        vals = [r[idx[col]] for r in data]
        missing = sum(1 for v in vals if v in (None, ""))
        dist = collections.Counter(str(v) for v in vals if v not in (None, ""))
        print(f"\n{col}  (missing/blank: {missing})")
        for k, v in dist.most_common():
            print(f"   {v:3}  {k}")

    problems = 0
    # duplicate ids
    ids = [str(r[idx["id"]]) for r in data]
    dup_ids = [k for k, c in collections.Counter(ids).items() if c > 1]
    if dup_ids:
        problems += len(dup_ids)
        print(f"\n⚠ DUPLICATE ids: {dup_ids}")
    # duplicate narratives
    narr = [str(r[idx["narrative"]]).strip() for r in data]
    dup_narr = [k for k, c in collections.Counter(narr).items() if c > 1 and k]
    if dup_narr:
        problems += len(dup_narr)
        print(
            f"\n⚠ DUPLICATE narratives: {len(dup_narr)} (first 3: {[n[:60] for n in dup_narr[:3]]})"
        )
    # out-of-vocab gold cells vs Option Sets
    oov: list[str] = []
    for r in data:
        for col in _GOLD:
            v = r[idx[col]]
            if v in (None, ""):
                continue
            if str(v).strip() not in allowed.get(col, set()):
                oov.append(f"id={r[idx['id']]} {col}={v!r}")
    if oov:
        problems += len(oov)
        print(f"\n⚠ OUT-OF-VOCAB gold cells ({len(oov)}):\n  " + "\n  ".join(oov[:20]))

    print("\n" + "=" * 60)
    print(
        "QA PASS — no integrity problems" if problems == 0 else f"QA: {problems} problem(s) above"
    )
    return 1 if problems else 0


def _parse(argv: list[str]) -> tuple[str, Path, bool]:
    """(name, workbook, do_qa). `--from <path>` overrides the default workbook; the first bare token is
    the label-source name (the CSV becomes holdout_labels_<name>.csv)."""
    name, workbook, do_qa = "owner", _WORKBOOK, False
    it = iter(argv)
    for a in it:
        if a == "--qa":
            do_qa = True
        elif a == "--from":
            workbook = Path(next(it))
        elif not a.startswith("-"):
            name = a
    return name, workbook, do_qa


def main() -> int:
    name, workbook, do_qa = _parse(sys.argv[1:])
    if not workbook.exists():
        print(f"no workbook at {workbook}")
        return 1
    if do_qa:
        return qa(workbook)
    export(name, workbook)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
